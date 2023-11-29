"""
@Time: 2023/11/28 14:34
@Auth: Y5neKO
@File: HighRiskVul.py
@IDE: PyCharm
"""
import json
import os.path
import re
import shutil
from urllib.parse import urlparse

import openpyxl
import requests
import xlrd
import win32com.client as win32


def xlsx2xls(book):
    """
    xlsx转xls，用于xlrd模块读取
    :param book:
    :return:
    """
    if not os.path.isabs(book):
        book = os.path.abspath(book)
    if not os.path.exists(book[:-1]):
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(book)
        wb.SaveAs(book[:-1], FileFormat=56)
        wb.Close(False)
        return book[:-1]
    else:
        return book[:-1]


def is_numeric(character):
    """
    数字判定
    :param character: 需要判定的字符串
    :return: 标识
    """
    flag = True
    for i1 in range(len(character)):
        if not character[i1].isdigit():
            flag = False
    return flag


def vul_type_iden(vul_name):
    """
    漏洞类型识别
    :param vul_name: 漏洞名称
    :return: 漏洞类型
    """
    vul_dict = {
        "SQL": "SQL注入",
        "XSS": "XSS",
        "跨站脚本": "XSS",
        "弱口令": "弱口令",
        "文件上传": "文件上传",
        "目录遍历": "目录遍历",
        "信息泄露": "信息泄露",
        "后门": "存在后门",
        "逻辑": "逻辑漏洞",
        "代码执行": "代码执行",
        "命令执行": "命令执行",
        "解析漏洞": "解析漏洞",
        "硬编码": "硬编码漏洞"
    }
    for key in vul_dict:
        if key in vul_name:
            return vul_dict[key]
    return "其他"


def ip_iden(ip, interface):
    """
    IP属地识别
    :param ip: ip地址
    :param interface: 接口名称
    :return: 省份、市区信息
    """
    result = {
        "pro": "未知",
        "city": "未知"
    }
    if interface == "太平洋":
        error_count = 0
        while True:
            try:
                print("正在向太平洋接口请求解析第" + str(error_count + 1) + "次")
                response = requests.get(
                    "http://whois.pconline.com.cn/ipJson.jsp?ip=" + str(ip) + "&json=true")
                data = json.loads(response.text)
                break
            except:
                error_count += 1
                print("解析失败")
                continue
        result = {
            "pro": data["pro"][:-1],
            "city": data["city"][:-1]
        }
    elif interface == "百度":
        error_count = 0
        while True:
            try:
                print("正在向百度接口请求解析第" + str(error_count + 1) + "次")
                response = requests.get("https://opendata.baidu.com/api.php?query=" + str(ip) + "&co=&resource_id=6006&oe=utf8")
                print(response.text)
                data = json.loads(response.text)
                location = data["data"][0]["location"]
                info, other = location.split(" ")[0:2]
                province, city = info.split("省")[0:2]
                break
            except:
                error_count += 1
                print("解析失败")
                continue
        result = {
            "pro": province,
            "city": city[:-1]
        }
    return result


def unit_iden(url):
    """
    单位识别
    :param url: 需要识别的url
    :return: 单位名称
    """
    response = requests.get(url)
    pattern = r'<a data-v-18f05699="" title="(.*?)" data-log-an="s-componylist-item-click" data-log-title="item-54225112166326">'
    matches = re.findall(pattern, response.text, re.DOTALL)


def write_xlsx(book, data_to_write, offset):
    """
    写入表格
    :param book: 表格路径
    :param data_to_write: 写入的数据
    :param offset: 写入的表格坐标
    :return: 无
    """
    workbook = openpyxl.load_workbook(book)
    worksheet = workbook["高危漏洞线索表"]
    # worksheet.parent.calc_on_load = False
    # worksheet.parent.enable_auto_filter = False
    worksheet[offset] = data_to_write
    workbook.save(book)


def write_xlsx_plus(book, data_to_write, offset):
    """
    写入表格plus
    :param book: 表格路径
    :param data_to_write: 写入的数据
    :param offset: 写入的表格坐标
    :return: 无
    """
    workbook = openpyxl.load_workbook(book)
    worksheet = workbook["高危漏洞线索表"]
    # 将所有数据存储在一个二维数组中
    data_array = []
    for row in range(len(data_to_write)):
        data_row = []
        for col in range(len(data_to_write[row])):
            cell_offset = chr(65 + col) + str(row + 2)
            data_row.append(data_to_write[row][col])
        data_array.append(data_row)
    # 批量写入数据
    for data_row in data_array:
        worksheet.append(data_row)
    workbook.save(book)


def vul_main(book, book_result):
    book_result_template = "template.xlsx"    # 模板位置
    if book.endswith(".xlsx"):
        book = xlsx2xls(book)
        print("-----------------Tips-----------------")
        print("导入表格为xlsx，为方便xlrd处理已自动转为xls，稍后会自动删除\n转换后地址：" + book)
        print("--------------------------------------")

    # 复制导出模板
    shutil.copy2(book_result_template, book_result)

    workbook = xlrd.open_workbook(book)
    # sheet_names = workbook.sheet_names()
    # print(sheet_names)
    worksheet = workbook.sheet_by_name("漏洞列表")
    i = 1
    # row_count = 1
    # print(worksheet.nrows)
    data_array = []
    for row_index in range(worksheet.nrows):
        details = []
        row_data = worksheet.row_values(row_index)
        # 如果第一个字段以数字开始，表示是漏洞详情行
        if is_numeric(str(row_data[0])):
            # row_count += 1
            if "高危" in str(row_data[1]) or "紧急" in str(row_data[1]):
                print("详情行数：" + str(row_index + 1))
                url = row_data[3]
                parsed_url = urlparse(url)
                # 解析ip地址
                # print("--------------------------------------")
                print("正在解析IP地址......")
                data = ip_iden(parsed_url.hostname, "百度")
                # while True:
                #     error_count = 0
                #     try:
                #         print("正在向接口请求解析第" + str(error_count + 1) + "次")
                #         response = requests.get(
                #             "http://whois.pconline.com.cn/ipJson.jsp?ip=" + str(parsed_url.hostname) + "&json=true")
                #         data = json.loads(response.text)
                #         break
                #     except:
                #         error_count += 1
                #         print("解析失败")
                #         continue
                print("解析完成")
                pro = data["pro"]
                city = data["city"]
                num = i
                details.append(num)
                details.append(parsed_url.netloc.split(":")[0])
                details.append(parsed_url.hostname)
                if parsed_url.port is None:
                    port = 80
                    if parsed_url.scheme == "https":
                        port = 443
                else:
                    port = parsed_url.port
                details.append(str(port))
                details.append(url)
                details.append(parsed_url.scheme)
                details.append(row_data[2])
                if row_data[6] != "":
                    vul_num = row_data[6]
                elif row_data[7] != "":
                    vul_num = row_data[7]
                elif row_data[8] != "":
                    vul_num = row_data[8]
                elif row_data[9] != "":
                    vul_num = row_data[9]
                else:
                    vul_num = "无"
                details.append(vul_num)
                details.append(vul_type_iden(row_data[2]))
                details.append(pro)
                details.append("")  # 发现时间
                details.append("")  # 是否异地
                details.append("")  # 单位
                details.append("暂无")  # 关联个人姓名
                details.append("暂无")  # 身份证
                details.append("暂无")  # 手机号
                details.append(pro)
                details.append(city)
                print("写入数据：\n" + str(details))
                # print("--------------------------------------")
                # print("正在写入......")
                # for col in range(65, 65 + len(details)):
                #     offset = str(chr(col)) + str(i + 2)
                #     # worksheet[offset] = details[col-65]
                #     write_xlsx(book_result, details[col - 65], offset)

                data_array.append(details)

                # workbook_result = openpyxl.load_workbook(book_result)
                # worksheet_result = workbook_result["高危漏洞线索表"]

                # print("写入完成")
                print("--------------------------------------")
                i += 1
        # print(row_count)  #打印有效漏洞详情行
    print(data_array)

    # 将最终得到的数据写入模板
    print("正在写入......")
    workbook_result = openpyxl.load_workbook(book_result)
    worksheet_result = workbook_result["高危漏洞线索表"]
    worksheet_result._current_row = 2
    for worksheet_result_row in range(len(data_array)):
        worksheet_result.append(data_array[worksheet_result_row])
    workbook_result.save(book_result)
    print("写入完成")

    os.remove(book)
    workbook.release_resources()
